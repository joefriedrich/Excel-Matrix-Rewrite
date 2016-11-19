'''
	Title:  Excel Matrix tool
	Author:  Joe Friedrich
	License:  MIT
'''

print('Importing libraries')
import openpyxl
import re
from sys import stdin
from warnings import simplefilter
from collections import namedtuple
from datetime import datetime
from tkinter import Tk

#------------------------Begin Company Class-----------------------------------------------------
class Company:
	def __init__(self, name, file_name, regular_expression, clients):
		self.name = name
		self.excel_file = openpyxl.load_workbook(file_name, read_only = True)
		self.role_format = regular_expression
		self.clients = clients
		self.vacation_lookup = self.load_vacations()
		self.region_lookup = self.load_regions()
		self.role_lookup = self.load_roles()
		self.email_lookup = self.load_emails()
		self.excel_file = None

		
	def load_regions(self):
		'''
			This grabs the contents of the cells in the top row of the Roles worksheet.
			It stores them in a list and stops when it gets to a blank cell.
			It returns all entries after the first 4.
			These are the headings that correspond with the companies regions.
		'''
		header_text = []
		for role in self.excel_file["Roles"]:
			for cell in role:
				if cell.value == None:
					break
				header_text.append(cell.value)
			break
		return header_text[4:]


	def vacation_check(self, names):
		'''
			Takes a list of strings (names)
			For each name in names, looks for name in vacation_lookup (list of tuples).
				If the name is there, checks the date against the dates in vacation_lookup.
				If both of these check out, it replaces name with the name in vacation_lookup.
			Returns a list of strings (checked_names)
		'''	
		checked_names = []
		for name in names:
			for vacation in self.vacation_lookup:
				if name == None:
					break
				elif name == vacation.standard:
					if vacation.start <= datetime.now() <= vacation.finish:
						name = vacation.replacement
						break
			checked_names.append(name)
		return checked_names


	def load_approvers(self, row):
		'''
			Takes the row and a blank approvers list.
			Approvers occur row[4] and beyond.
			Fills the approvers list with all approvers in the row 
				when row[2] is Regional.
			Otherwise, only grabs the first approver it finds.
		'''
		approvers = []
		if row[2].value == 'Regional':
			approvers = [approver.value for approver in row[4:4 + len(self.region_lookup)]]
		else:
			for approver in row[4:4 + len(self.region_lookup)]:
				if approver.value != 'N/A':
					approvers.append(approver.value)
					break
		approvers = self.vacation_check(approvers)
		return approvers

	
	def load_roles(self):
		'''
			This creates a list of tuples that contain
				the role name.
				the role definition.
				and a list of the role approvers (created by load_approvers).
			Region was not included, because we can get that from
				the length of the list of role approvers (when greater than 1).
		'''
		print('Loading lookup dictionary...')
		role = namedtuple('role',
						  'name description approvers')
		role_list = []
		for row in self.excel_file['Roles'].rows:
			if row[0].value in ('', None):
				print('Issue with a row.')   #Can we pull the excel row number from this?
				break
			role_tuple = role(row[0].value, row[1].value, self.load_approvers(row))
			role_list.append(role_tuple)
		return role_list


	def load_emails(self):
		'''
			Creates a dictionary of emails from the table of names/emails in the sprreadsheet.
		'''
		emails = {}
		for row in self.excel_file['Names and Email'].rows:
			emails[row[0].value] = row[1].value
		return emails


	def load_vacations(self):
		'''
			Creates a list of tuples from the information in the OnHoliday spreadsheet.
		'''
		vacation = namedtuple('vacation',
				'standard replacement start finish')
		vacation_list = []
		for row in self.excel_file['OnHoliday'].rows:
			vacation_tuple = vacation(row[0].value,
						row[1].value,
						row[2].value,
						row[3].value)
			vacation_list.append(vacation_tuple)
		return vacation_list


	def find_and_sort_roles(self, user_input, user_region):
		'''
			Takes a list of strings (user_input) and an int user_region.
			It filters each string of user_input through that companies regex.
			[It forces the input string to be all uppercase letters.]
			Takes the results of that filter and 
				tries to find it in that company's role_lookup.
			If it is a Regional role, the lenght of appovers will be > 1.
				Collects tuples (name, description, approver).
			Returns the list of tuples sorted by approver name.
		'''
		roles_and_approvers = []
		for line in user_input:
			find_role = self.role_format.search(line.upper())
			role_not_found = True
			if(find_role != None):
				for row in self.role_lookup:
					if find_role.group() == row.name:
						if len(row.approvers) > 1:
							roles_and_approvers.append((row.name,
											row.description, 
											row.approvers[user_region]))
						else:
							roles_and_approvers.append((row.name, 
											row.description, 
											row.approvers[0]))
						role_not_found = False
				if role_not_found:
					print('\nRole ' + find_role.group() + ' was not found.')
		return sorted(roles_and_approvers, key = lambda entry: entry[2])

		
#---------------------------End Company Class---------------------------------------------------
def get_menu(list_of_things):
	'''
		This takes a list as input.
		It prints the list to the screen as menu options 1 through len(list).
		It calls and returns the results of get_menu_input.
	'''
	menu_number = 0
	for thing in list_of_things:
		menu_number += 1
		print(str(menu_number) +  ':  ' + thing)
	return get_menu_input(menu_number)


def get_menu_input(menu_total):
	'''
		This takes the final number of possible selections (whole number).
		It takes user input and
			-if the input is between 1 and the last possible selection
				it returns the number the user typed in.
			-if the number is greater than the last possible selection
				or less than 1
				or not a whole number
				or not a number at all
				it returns 0.
	'''
	while True:
		try:
			user_input = int(input('Type a number from the menu and hit enter.\nMake any other entry and enter to stop:  '))
			if 1 <= user_input <= menu_total:
				return user_input
			else:  #the numerical value is not a menu option
				quitting = input('\nAre you sure you are finished?\nPress y and enter to quit.\nAny other key(s) and enter to continue:  ')
				if quitting == 'y':
					return 0
		except ValueError:  #if the value is a non-numeral or float
			quitting = input('\nAre you sure you are finished?\nPress y and enter to quit.\nAny other key(s) and enter to continue:  ')
			if quitting == 'y':
				return 0

				
def get_role_input():
	'''
		This is just a placeholder for user input that is not in the form
			of get_menu_input.
	'''
	return stdin.readlines()

	
def parse_role_tuples(output, company, clipboard):
	'''
		Takes the list of tuples (output), the company object, and the clipboard.
			Output is expected to be sorted by approver (output[2]).
			Clipboard should be blank.
		For every unique approver in the tuples, it will print and append a header
			to the clipboard.
			It will also append the email address to approver_emails.
		These are followed by the role names until a new approver name is found or
			the list ends.
		Returns a list of strings (approver_emails) and the clipboard.
	'''
	current_approver = ''
	approver_emails = []
	for role_tuple in output:
		if(role_tuple[2] != current_approver):
			current_approver = role_tuple[2]
			print('\n' + user_client + ' -- awaiting approval from ' + current_approver)
			clipboard.clipboard_append('\n' + user_client + ' -- awaiting approval from ' + current_approver + '\n')
			if current_approver in company.email_lookup:
				approver_emails.append(company.email_lookup[current_approver])
			else:
				approver_emails.append(current_approver + "'s email is missing")
		print(role_tuple[0] + '\t ' + role_tuple[1])
		clipboard.clipboard_append(role_tuple[0] + '\t ' + role_tuple[1] + '\n')
	return approver_emails, clipboard

	
def single_approvers(single_approver_data, email_list, clipboard):
	'''
		Takes list of tuples, list of strings, and clipboard.
		Will take one single approver at a time and append it's data to 
			email_list and clipboard.
		Returns list of strings (email_list) and clipboard.
	'''
	while True:
		print('\n************Single Approver Client Section************************')
		print('***This will print your single client approvers UNTIL ************')
		print('*****you make a selection that is NOT on the menu.****************')
		print('******************************************************************')
		print('\nDoes the user need additional access in single approver clients?')
		select_single_client = get_menu(single_approver_data)
		if select_single_client == 0:
			break
		
		selected_client = single_approver_clients[select_single_client - 1]
		
		print(selected_client[1])
		clipboard.clipboard_append(selected_client[1])
		if selected_client[2] not in email_list:
			email_list.append(selected_client[2])
		
	return email_list, clipboard

	
def email_format(email_list):
	'''
		Takes a list of strings.
		Returns a concatination of this list.
	'''
	email_output = ''
	for email in email_list:
		email_output += email + ', '	

	print('\n' + email_output[:-2] + '\n')
	return str('\n' + email_output[:-2] + '\n')
	

#--------------------------------End Local Functions-----------------------------------------------
simplefilter('ignore')  #blocks some warnings that openpyxl throws when loading files

print('Loading companies.')
company1 = Company('Company1',
			openpyxl.load_workbook(r'/{file path}/matrixCompany1.xlsx'),
			re.compile(r'[A-Z]{1,3}(:|_)\S+'),
			['ProdC1', 'QaC1', 'ProdC1/QaC1', 'DevC1', 'QaC1/DevC1'])
company2 = Company('Company2',
			openpyxl.load_workbook(r'/{file path}/matrixCompany2.xlsx'),
			re.compile(r'Z:\S{4}:\S{7}:\S{4}:\S'),
			['ProdC2', 'QaC2', 'ProdC2/QaC2', 'DevC2', 'QaC1/DevC2'])
company3 = Company('Company3',
			openpyxl.load_workbook(r'/{file path}/matrixCompany3.xlsx'),
			re.compile(r'\S+'),
			['ProdC3', 'QaC3', 'ProdC3/QaC3', 'DevC3', 'QaC3/DevC3'])

list_companies = [company1, company2, company3]

print('Loading single approvers.')		
single_approver_clients = []
single_approver_clients.append(('Company One Special Test Client', 'CP1TST -- awaiting approval from singleApproverCP1/nSpecialRole', 'singleApprover@company1.com'))
single_approver_clients.append(('Company Two Special Test Client', 'CP2TST -- awaiting approval from singleApproverCP2/nSpecialRole', 'singleApprover@company2.com'))
single_approver_clients.append(('Company Three Special Test Client', 'CP3TST -- awaiting approval from singleApproverCP3/nSpecialRole', 'singleApprover@company3.com'))
single_approver_clients.append(('Company One Test Email Duplication', 'CP1TST -- awaiting approval from ApproverOneCP1/nSpecialRole', 'Approver1@company1.com'))
single_approver_clients.append(('Company Two Test Email Duplication', 'CP2TST -- awaiting approval from ApproverOneCP2/nSpecialRole', 'Approver1@company2.com'))
single_approver_clients.append(('Company Three Test Email Duplication', 'CP3TST -- awaiting approval from ApproverOneCP3/nSpecialRole', 'Approver1@company3.com'))

list_company_names = [company.name for company in list_companies]
list_single_approvers = [client[0] for client in single_approver_clients]

clipboard = Tk()
clipboard.withdraw() #removes Tk window that is not needed at this time.
			
#--------------------------------Begin Program-----------------------------------------------
while (True):	
	print('\nTo which company does the user belong?')
	select_company = get_menu(list_company_names)
	if select_company == 0:
		clipboard.destroy()  
		break
	company = list_companies[select_company - 1]
	
	print('\nWhich region does the user wish to access?')
	select_region = get_menu(company.region_lookup)
	if select_region == 0:
		clipboard.destroy()  
		break
	region = select_region - 1
	
	print("\nPaste the roles in, hit Ctrl+D (Ctrl+Z and Enter for Windows).")
	requested_roles = get_role_input()
	
	role_tuples = company.find_and_sort_roles(requested_roles, region)
	
	print('\nIn which SAP client does the user want the access?')
	select_client = get_menu(company.clients)
	if select_client == 0:
		clipboard.destroy()  
		break
	user_client = company.clients[select_client - 1]
	
	clipboard.clipboard_clear() #clear's the current contents of the clipboard in prep for new data
	email_list, clipboard = parse_role_tuples(role_tuples, company, clipboard)
	
	email_list, clipboard = single_approvers(list_single_approvers, email_list, clipboard)
	
	clipboard.clipboard_append(email_format(email_list))
	print('\nYour output is now in the clipboard.  Paste it into your ticket.')
